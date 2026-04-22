"""
Servicio de exportación a Excel.
"""
import logging
import os
from datetime import datetime

from openpyxl import Workbook

logger = logging.getLogger(__name__)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
EXITO_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FALLO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITULO_FONT = Font(bold=True, size=14, color="4472C4")
LABEL_FONT = Font(bold=True, size=11)
BORDE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class ExportService:
    """Exporta resultados de ejecución a Excel."""

    def export_excel(
        self,
        registros: list,
        resumen: dict,
        estructura: dict,
        export_dir: str = "exports",
    ) -> str:
        """Exporta los datos a Excel.

        Returns:
            Ruta absoluta del archivo Excel generado.
        """
        os.makedirs(export_dir, exist_ok=True)

        wb = Workbook()

        # ===== HOJA 1: RESPUESTAS =====
        ws = wb.active
        ws.title = "Respuestas"

        preguntas_headers = []
        for pagina in estructura.get("paginas", []):
            for preg in pagina.get("preguntas", []):
                if preg["tipo"] != "informativo":
                    preguntas_headers.append(preg["texto"])

        headers = ["#", "Estado", "Tiempo (s)", "Perfil", "Tendencia"] + preguntas_headers
        self._escribir_headers(ws, headers)

        for fila, reg in enumerate(registros, 2):
            datos = [
                reg["numero"],
                "Exitosa" if reg["exito"] else "Fallida",
                f"{reg['tiempo']:.1f}",
                reg.get("perfil", ""),
                reg.get("tendencia", ""),
            ]

            valores = {}
            for pagina in reg["respuestas"].get("paginas", []):
                for r in pagina.get("respuestas", []):
                    valores[r["pregunta"]] = r["valor"]

            for preg in preguntas_headers:
                valor = valores.get(preg, "")
                datos.append(self._excel_value(valor))

            self._escribir_fila(ws, fila, datos, col_estado=2)

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18

        # ===== HOJA 2: RESUMEN =====
        ws2 = wb.create_sheet("Resumen")
        ws2.cell(row=1, column=1, value="RESUMEN DE EJECUCIÓN").font = TITULO_FONT
        ws2.merge_cells("A1:C1")

        datos_resumen = [
            ("Fecha", resumen["fecha"]),
            ("Total", resumen["total"]),
            ("Exitosas", resumen["exitosas"]),
            ("Fallidas", resumen["fallidas"]),
            ("Tasa de éxito", f"{resumen['tasa_exito']:.1f}%"),
            ("Tiempo total", resumen["tiempo_total_fmt"]),
            ("Tiempo promedio", resumen["tiempo_promedio_fmt"]),
        ]
        for i, (label, valor) in enumerate(datos_resumen, 3):
            ws2.cell(row=i, column=1, value=label).font = LABEL_FONT
            ws2.cell(row=i, column=2, value=valor)

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 25

        # Guardar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"encuestas_{timestamp}.xlsx"
        ruta = os.path.join(export_dir, nombre)
        wb.save(ruta)
        logger.info("Excel exportado: %s", nombre)

        return os.path.abspath(ruta)

    def _escribir_headers(self, ws, headers, fila=1):
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col, value=header)
            celda.font = HEADER_FONT
            celda.fill = HEADER_FILL
            celda.alignment = Alignment(horizontal="center", wrap_text=True)
            celda.border = BORDE

    def _escribir_fila(self, ws, fila, datos, col_estado=None):
        for col, valor in enumerate(datos, 1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.border = BORDE
            celda.alignment = Alignment(horizontal="center")
            if col_estado and col == col_estado:
                celda.fill = (
                    EXITO_FILL
                    if "xitos" in str(valor).lower() or "exitosa" in str(valor).lower()
                    else FALLO_FILL
                )

    @staticmethod
    def _excel_value(valor):
        if isinstance(valor, dict):
            partes = [f"{str(k).strip()}: {str(v).strip()}" for k, v in valor.items()]
            return " | ".join(partes)
        if isinstance(valor, list):
            return ", ".join(str(v) for v in valor)
        return valor
